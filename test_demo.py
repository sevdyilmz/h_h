from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.action_chains import ActionChains
import pytest
from pathlib import Path
from datetime import date
import openpyxl
from constants import globalConstants
#prefix= ön ek test_
#fostfix 

class Test_DemoClass:
    #her testten önce çağırılır. ansayfa için kullanılabilir.
    def setup_method(self):
        self.driver=webdriver.Chrome()
        self.driver.maximize_window()
        self.driver.get(globalConstants.URL)
        #günün tarihin al bu tarihli dosya var mı kondtol et yoksa oluştur.
        # today=date.today()
        # print(today)
        self.folderPath=str(date.today()) #==>12.04.2024
        Path(self.folderPath).mkdir(exist_ok=True)

    #her testten sonra çağırılır. her test sonrası chromu kapat gibi kullanılabili.r
    def teardown_method(self):
        self.driver.quit()

    def test_demoFunc(self):
        text="hello"
        assert text =="hello"
    def test_demo2(self):
        assert True
    
    def waitForElementVisible(self, locator):
        WebDriverWait(self.driver,5).until(ec.visibility_of_element_located(locator))
        WebDriverWait(self.driver,5).until(ec.visibility_of_element_located(locator))
        
    # @pytest.mark.skip()
    # def test_invalid_login(self):
    #     usernameInput=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,"user-name")))
    #     passwordInput=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID, "password")))
    #     usernameInput.send_keys("1")
    #     passwordInput.send_keys("1")
    #     loginBtn=self.driver.find_element(By.XPATH,"//*[@id='login-button']")
    #     sleep(2) 
    #     loginBtn.click()
    #     sleep(2)
    #     errorMessage=self.driver.find_element(By.XPATH,"//*[@id='login_button_container']/div/form/div[3]/h3")
    #     assert errorMessage.text=="Epic sadface: Username and password do not match any user in this service"
    
    def getData():
        #veriyi al
        excelFile=openpyxl.load_workbook("data/invalid_login.xlsx")
        selectedSheet=excelFile["Sayfa1"]
        totalRows=selectedSheet.max_row
        data=[]
        for i in range(2, totalRows+1):

            username=selectedSheet.cell(i,1).value
            password=selectedSheet.cell(i,2).value
            tupleData=(username,password)
            data.append(tupleData)
         
        return data

    @pytest.mark.parametrize("username, password", getData())
    def test_invalid_login(self, username, password):
        self.waitForElementVisible((By.ID,"user-name"))
        usernameInput=self.driver.find_element(By.ID,"user-name")
        self.waitForElementVisible((By.ID, "password"))
        passwordInput=self.driver.find_element(By.ID, "password")
        usernameInput.send_keys(username)
        passwordInput.send_keys(password)
        loginBtn=self.driver.find_element(By.XPATH,"//*[@id='login-button']")
        sleep(2) 
        loginBtn.click()
        sleep(2)
        errorMessage=self.driver.find_element(By.XPATH,"//*[@id='login_button_container']/div/form/div[3]/h3")
        self.driver.save_screenshot(f"{self.folderPath}/test_invalid_login-{username}-{password}.png")
        #tırnaklar arası uyarı mesajına normalde magic string deniyor. ama bu genelde tercih edilmiyor çünkü birkaç yerde aynı mesaj kullanılıyor olabilir ve tek bir işaret bile unutulsa hata alınabilir ve farkedilmesi zor olur. 
        assert errorMessage.text=="Epic sadface: Username and password do not match any user in this service"
    
